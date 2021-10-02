import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from scipy import stats

import pycurious

# load magnetic anomaly - i.e. random fractal noise
mag_data=np.loadtxt("C:\\Users\\Izzul Qudsi\\Downloads\\MOD 14-15_Proposal\\Mijn Thesis\\CPDbouligand\\400_edit.txt")
#nx,ny= 3662,2789

x = mag_data[:,0]
y = mag_data[:,1]

xmin, xmax = x.min(), x.max()
ymin, ymax = y.min(), y.max()
print (xmax, xmin, ymin, ymax)

dx, dy = 400, 400 # 400 m grid resolution
nx, ny = int(round((xmax-xmin)/dx)+1), int(round((ymax-ymin)/dy)+1)

d = mag_data[:,2].reshape(ny,nx)

# filter NaNs
mag_data = mag_data[mag_data[:,2] != 9999.]

grid = pycurious.CurieOptimise(d, xmin, xmax, ymin, ymax)

# get centroids

window_size = 400e3
xc_list, yc_list = grid.create_centroid_list(window_size, spacingX=50e3, spacingY=50e3)

print("number of centroids = {}".format(len(xc_list)))
#print (xc_list)
#print (yc_list)

# no priors
grid.reset_priors()
grid.add_prior(zt=(5.0,0.1), dz=(35.0,10.0))

beta, zt, dz, C = grid.optimise_routine(window_size, xc_list, yc_list)

#tes plot
#window_size = 400e3

#subgrid = grid.subgrid(window_size, xc_list, yc_list)
#Phi2 = pycurious.bouligand2009(k, beta, zt, dz, C)
#k, Phi, sigma_Phi = grid.radial_spectrum(subgrid)
#plt.plot(k, Phi, '-o')
#plt.plot(k, Phi2, linewidth=2)

# get dimensions of domain
xcoords = np.unique(xc_list)
ycoords = np.unique(yc_list)
nc, nr = xcoords.size, ycoords.size
#print (xcoords)
#print (ycoords)
print (nc,nr)

# plot results
fig, (ax1, ax2, ax3, ax4) = plt.subplots(1, 4, sharex=True, sharey=True, figsize=(17,3.))

im1 = ax1.imshow(beta.reshape(nr,nc))
im2 = ax2.imshow(zt.reshape(nr,nc))
im3 = ax3.imshow(dz.reshape(nr,nc))
im4 = ax4.imshow(C.reshape(nr,nc))

fig.colorbar(im1, ax=ax1, label=r"$\beta$")
fig.colorbar(im2, ax=ax2, label=r"$z_t$")
fig.colorbar(im3, ax=ax3, label=r"$\Delta z$")
fig.colorbar(im4, ax=ax4, label=r"$C$")

np.savetxt("beta.txt", beta, delimiter=" ")
np.savetxt("zt.txt", zt, delimiter=" ")
np.savetxt("dz.txt", dz, delimiter=" ")
np.savetxt("C.txt", C, delimiter=" ")

# plot Curie depth

#curie_depth = zt + dz

#fig = plt.figure()
#ax1 = fig.add_subplot(111)
#im1 = ax1.imshow(curie_depth.reshape(nr,nc), cmap=plt.cm.BrBG)
#fig.colorbar(im1)

#scale of each parameter (check again for the search range in the code)!!!!!!!!!!!!!!!
x_scale = [0.25, 0.1, 1.0, 0.5]

# run more simulations for production runs
burnin = 1000
nsim = 5000

# mean across the domain
mu_beta, mu_zt, mu_dz, mu_C = beta.mean(), zt.mean(), dz.mean(), C.mean()

pt_post = []

# This will take some time
for xc, yc in zip(xc_list, yc_list):
    posterior = grid.metropolis_hastings(window_size, xc, yc, nsim, burnin, x_scale,\
                                         mu_beta, mu_zt, mu_dz, mu_C, taper=None)
    pt_post.append( posterior )

curie_depth = np.zeros_like(xc_list)
uncertainty = np.zeros_like(xc_list)

for i, pt in enumerate(pt_post):
    betaP, ztP, dzP, CP = pt
    cpd = ztP + dzP
    curie_depth[i] = np.mean(cpd)
    uncertainty[i] = np.std(cpd) #standard deviation of the result

# plot Curie depth

#curie_depth = zt + dz

#fig, (ax1,ax2) = plt.subplots(1,2, figsize=(11,4))
#im1 = ax1.imshow(curie_depth.reshape(nr,nc), cmap=plt.cm.BrBG)
#im2 = ax2.imshow(uncertainty.reshape(nr,nc), cmap=plt.cm.Blues)
#fig.colorbar(im1, ax=ax1)
#fig.colorbar(im2, ax=ax2)

#ax1.set_title("Curie depth (km)")
#ax2.set_title("Uncertainty (km)")

#sensitivity

beta_p = stats.norm(3.0, 1.0)
grid.add_prior(beta=beta_p)

nsim = 100
pt_post = []

for xc, yc in zip(xc_list, yc_list):
    sensitivity = grid.sensitivity(window_size, xc, yc, nsim, mu_beta, mu_zt, mu_dz, mu_C, taper=None)
    pt_post.append( sensitivity )

curie_depth = np.zeros_like(xc_list)
uncertainty = np.zeros_like(xc_list)

for i, pt in enumerate(pt_post):
    betaP, ztP, dzP, CP = pt
    cpd = ztP + dzP
    curie_depth[i] = np.mean(cpd)
    uncertainty[i] = np.std(cpd)

# plot Curie depth

curie_depth = zt + dz

fig, (ax1,ax2) = plt.subplots(1,2, figsize=(11,4))
im1 = ax1.imshow(curie_depth.reshape(nr,nc), cmap=plt.cm.BrBG)
im2 = ax2.imshow(uncertainty.reshape(nr,nc), cmap=plt.cm.Blues)
fig.colorbar(im1, ax=ax1)
fig.colorbar(im2, ax=ax2)
#print (curie_depth)
ax1.set_title("Curie depth (km)")
ax2.set_title("Uncertainty (km)")

np.savetxt("xcoord.txt", xc_list, delimiter=" ")
np.savetxt("ycoord.txt", yc_list, delimiter=" ")
np.savetxt("CPD.txt", curie_depth, delimiter=" ")
np.savetxt("uncertainty.txt", uncertainty, delimiter=" ")

#to import the result to excel sheets
def to_excel():
    path= 'C:\\Users\\Izzul Qudsi\\Downloads\\MOD 14-15_Proposal\\Mijn Thesis\\CPDbouligand\\Scripts\\toexceltest.xlsx'
    wb=load_workbook(path)
    a=nc
    b=nr
    with open ('C:\\Users\\Izzul Qudsi\\Downloads\\MOD 14-15_Proposal\\Mijn Thesis\\CPDbouligand\\xcoord.txt', 'r+') as d:
        xcoords=d.readlines()
    with open ('C:\\Users\\Izzul Qudsi\\Downloads\\MOD 14-15_Proposal\\Mijn Thesis\\CPDbouligand\\ycoord.txt', 'r+') as e:
        ycoords=e.readlines()
    with open ('C:\\Users\\Izzul Qudsi\\Downloads\\MOD 14-15_Proposal\\Mijn Thesis\\CPDbouligand\\CPD.txt', 'r+') as f:
        cpd=f.readlines()
    with open ('C:\\Users\\Izzul Qudsi\\Downloads\\MOD 14-15_Proposal\\Mijn Thesis\\CPDbouligand\\uncertainty.txt', 'r+') as g:
        uncertainty=g.readlines()
    z=1
    for z in range (1,b+1):
    #for z in range (1,4):
        ws=wb[('Sheet'+str(z))]
        #for row in range (1,42):
        for row in range (1,a+1):
            xcoor=ws.cell(row=row, column=1)
            xcoor.value=float(xcoords[0])
            del xcoords[0]
            ycoor=ws.cell(row=row, column=2)
            ycoor.value=float (ycoords[0])
            del ycoords[0]
            curie=ws.cell(row=row, column=3)
            curie.value=float (cpd[0])
            del cpd[0]
            uncer=ws.cell(row=row, column=4)
            uncer.value=float (uncertainty[0])
            del uncertainty[0]
        wb.create_sheet('Sheet'+str(z+1))
        spath=r'C:\Users\Izzul Qudsi\AppData\Local\Programs\Python\Python36\Scripts\result\''
        final=spath+('CPD400_test.xlsx')
        wb.save(final)
        
#import to excel and show the image
to_excel()
plt.show()
